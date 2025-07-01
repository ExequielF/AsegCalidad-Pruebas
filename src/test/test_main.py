import pytest
from src.main import ValFecha, ValCorreo, ValContrasena, ValUsuario
import pandas as pd
import openpyxl

#pip install pandas openpyxl pytest

excel = openpyxl.load_workbook("Datos/datos_anonimizados.xlsx")
datos = excel["Sheet1"]
totalRows = datos.max_row
totalCols = datos.max_column

mainlist = []
for i in range(2, totalRows+1):
    datalist = []
    for j in range(1, totalCols+1):
        data = datos.cell(row=i, column=j).value
        datalist.insert(j-1,data)
    mainlist.insert(i,datalist)

#############################################

primeras = mainlist[0:5]

conjunto_fechas = [(int(fila[4]), int(fila[5]), True) for fila in primeras] 
#Eva 4
@pytest.mark.parametrize(
    "mes, dia, esperado",
    [
        (8,18, True),
        (0,30, False),
        (13,20, False),
        (-1,4, False),
        (7,33, False),
        (10,"asas", False),
        (1,21, True),
    ]+ conjunto_fechas
)

def test_parametrize(mes, dia, esperado):
    assert ValFecha(mes, dia) == esperado

def test_fecha_valido():
    assert ValFecha(9, 14) is True

def test_mes_invalido():
    assert ValFecha(15, 27) is False

def test_mes_invalido1():
    assert ValFecha(0, 5) is False

def test_mes_invalido2():
    assert ValFecha(-1, 50) is False

def test_dia_invalido():
    assert ValFecha(4, 50) is False

def test_texto_invalido2():
    assert ValFecha(9, "abc") is False
    
def test_fecha_expirada():
    assert ValFecha(5, 21) is True

#################

conjunto_correos = [(fila[6], True) for fila in primeras] 

@pytest.mark.parametrize(
    "correo, esperado",
    [
        (mainlist[1][6], True), #Funciona
        ("angela89@gmail.com", True),
        ("aaaa_a.com", False),
        ("aaaa@a@a.com", False),
        (123, False),
        ("aaa@inacapmail.com", False),
        ("yrichardson@yahoo.com", True),
    ]+ conjunto_correos
)

def test_parametrize2(correo, esperado):
    assert ValCorreo(correo) == esperado

def test_correo_valido():
    assert ValCorreo("hola@hotmail.com") is True

def test_correo_arroba():
    assert ValCorreo("aaaa_a.com") is False

def test_correo_arroba1():
    assert ValCorreo("aaaa@a@a.com") is False

def test_correo_numero():
    assert ValCorreo(123) is False

def test_correo_dominio():
    assert ValCorreo("aaa@inacapmail.com") is False

#################

conjunto_contrasena = [(fila[3], True) for fila in primeras] 

@pytest.mark.parametrize(
    "contrasena, esperado",
    [
        (mainlist[1][3], True), #Funciona, pero las contraseñas no son validas ya que no cumple con la regla minima
        ("Abcdefghijklmnopq123###",True),
        ('ASDASDasda""""asdasd"12332',True),
        ("123",False),
        (True, False),
        ("1", False),
        ("adfadfasdfasdfasdfasdfasdfdafasdfasdfasfasf", False),
        ("12311231312312313", False),
        ("ASDASDADADASDASDASDASD", False),
        ("asdasdadadasdasdasd", False),
        ("AAAAAAAA1234567", False),
        ("aaaaaaaaaa1234567", False),
        ("aaaaaaaaAAAAAAAAAA", False),
    ]+ conjunto_contrasena
)

def test_parametrize3(contrasena, esperado):
    assert ValContrasena(contrasena) == esperado


def test_contrasena_con_valido():
    assert ValContrasena('ASDASDasdasdasd123123#$$#') is True

def test_contrasena_val_con_comillas():
    assert ValContrasena('ASDASDasda""""asdasd"12332') is True

def test_contrasena_entero():
    assert ValContrasena(123) is False

def test_contrasena_booleano():
    assert ValContrasena(True) is False

def test_contrasena_corta():
    assert ValContrasena("1") is False

def test_contrasena_larga():
    assert ValContrasena("adfadfasdfasdfasdfasdfasdfdafasdfasdfasfasf") is False

def test_contrasena_numeros():
    assert ValContrasena("12311231312312313") is False
    
def test_contrasena_mayusculas():
    assert ValContrasena("ASDASDADADASDASDASDASD") is False

def test_contrasena_minusulas():
    assert ValContrasena("asdasdadadasdasdasd") is False

def test_contrasena_may_num():
    assert ValContrasena("AAAAAAAA1234567") is False

def test_contrasena_min_num():
    assert ValContrasena("aaaaaaaaaa1234567") is False

def test_contrasena_min_may():
    assert ValContrasena("aaaaaaaaAAAAAAAAAA") is False


#############


@pytest.mark.parametrize(
    "usuario, contraseña, esperado",
    [
        ("usuario1", "Amarillo123", True), 
        ("usuario1", "clave456", False),
        ("usuario_no_existente", "contraseña123", False),
        ("usuario1", "contraseña_incorrecta", False),
    ]
)

def test_parametrize4(usuario, contraseña, esperado):
    assert ValUsuario(usuario, contraseña) == esperado


def test_login_usuario_correcto():
    assert ValUsuario("usuario1", "Amarillo123") is True

def test_login_usuario_incorrecto():
    assert ValUsuario("usuario1", "durAzno456") is False

def test_login_usuario_no_existente():
    assert ValUsuario("usuario_no_existente", "Amarillo123") is False

def test_login_contraseña_incorrecta():
    assert ValUsuario("usuario1", "contraseña_incorrecta") is False

def test_login_inyeccion():
    assert ValUsuario("'OR '1'='1", "'OR '1'='1") is False